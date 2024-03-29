#!/usr/bin/python3
import json
from string import  ascii_uppercase as ALPH

import telebot
from collections import Counter
from flask import Flask, g, redirect, render_template, request, url_for, session, send_file, jsonify
import sqlite3
from openpyxl import Workbook, load_workbook
from io import  BytesIO
from  datetime import date, timedelta
from config import DB_PATH, HOST, PORT, LOGIN, PAS, O_LOGIN, TELEGRAM_TOKEN, TRUSTED
from threading import Thread
import bot
from base64 import b64encode as b64enc, b64decode as b64dec

week_days = {1:"Понедельник",2:"Вторник",3:"Среда",4:"Четверг",5:"Пятница",6:"Суббота",7:"Воскресенье",}
ALPH = list(ALPH)
ALPH += ["AA", "AB", "AC", "AD", "AE", "AF", "AG", "AH", "AI", "AK", "AL", "AN"]

app = Flask(__name__)
app.config.from_object('config')

h1 = '<h1>{}</h1>'
cap = lambda s: " ".join(list(map(lambda st: st.capitalize(), s.split())))
up = lambda s : s.strip().upper()

form = lambda key: request.form[key]
form_get = lambda key, ret: request.form.get(key, ret)

def p(*items):
    print('\n-----------------------\n')
    for i in items:
        print('\t', i)
    print('\n-----------------------\n')


tele_bot = telebot.TeleBot(TELEGRAM_TOKEN)

@tele_bot.message_handler(content_types=['text'])
def get_text_messages(message):
    uid = message.from_user.id
    if uid in TRUSTED:
        bot.send_files(uid, tele_bot)
    else:
        tele_bot.send_message(uid, "Ты чужой, почему твой UID = %d?"%uid)


def get_db():
    db = getattr(g, 'db', None)
    if db is None:
        db = g.db = sqlite3.connect(DB_PATH)
    return db


@app.teardown_appcontext
def close_connection(_):
    db = getattr(g, 'db', None)
    if db is not None:
        db.close()


@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':

        lgn = request.form.get('login', None)
        if not lgn: h1.format("Не был получен логин пользователя")
        lgn = lgn.lower()
        if (lgn == LOGIN or lgn == O_LOGIN) and request.form.get('password', None) == PAS:
            session['login'] = lgn
            return redirect(url_for('index'))

        return redirect(url_for('login'))

    else:
        return render_template('login.html')


def make_xlsx(stream, when="cur"):
    xlsx = Workbook()
    db = sqlite3.connect(DB_PATH)
    cur = db.cursor()

    if when != 'cur':
        today = date.today()
        first = today.replace(day=1)
        d = first - timedelta(days=1)
    else:
        d = date.today()

    sheet = xlsx['Sheet']
    xlsx.remove(sheet)

    sheet = xlsx.create_sheet('Даты')
    dates = cur.execute('SELECT d, m, y, uid FROM proc WHERE  m=? AND y=? ORDER BY uid', (d.month, d.year)).fetchall()
    date_dict = {}
    for i in dates:
        dt = (i[2], i[1], i[0])
        if date_dict.get(dt) is None: date_dict[dt] = []
        date_dict[dt].append(i[3])
    sorted_dates = sorted(date_dict.items(), key=lambda x: x[0][2])
    for i in sorted_dates:
        sheet.append(['%d.%d.%d' % (i[0][2], i[0][1], i[0][0]), len(i[1])] + i[1])

    sheet = xlsx.create_sheet('Процедуры')
    procs = cur.execute('SELECT * FROM proc WHERE  m=? AND y=? ORDER BY uid', (d.month, d.year)).fetchall()
    sheet.append(['ФИО', 'Класс', 'Сделано процедур'])
    proc = {}
    for i in procs:
        if proc.get(i[0]) is None: proc[i[0]] = []
        proc[i[0]].append('%d.%d.%d' % (i[1], i[2], i[3]))

    for i in proc.items():
        fam, grade = i[0].split(' — ')[:2]
        fam = cap(fam)
        sheet.append([fam, grade, len(i[1])] + sorted(i[1], key=lambda x: int(x.split('.')[0])))

    xlsx.save(stream)
    xlsx.close()
    stream.seek(0)
    return stream.getvalue()


@app.route('/download/<when>')
def download(when='cur'):
    if not session.get('login'): return redirect(url_for('login'))
    stream = BytesIO()
    if session["login"] == LOGIN:
        make_xlsx(stream, when)
        return send_file(stream, cache_timeout=0, as_attachment=True, attachment_filename='massage_counter.xlsx')
    else:
        o_make_xlsx(stream, when)
        return send_file(stream, cache_timeout=0, as_attachment=True, attachment_filename='auto-report.xlsx')


@app.route('/addNew', methods=['POST'])
def addNew():
    if not session.get('login'): return redirect(url_for('login'))
    db = get_db()
    cur = db.cursor()
    if session["login"] == LOGIN:
        if form_get('fam', None) is None or form_get('grade', None) is None or form_get('fam', None) == '' or form_get('grade', None) == '': return h1.format('Не получено имя или класс ученика')
        cur.execute('INSERT INTO people VALUES (?,0,?)', (cap(form('fam'))+' — '+up(form('grade')),None))
    else:
        if request.files.get('file') is None: return h1.format('Не загружен файл')
        file = request.files['file']
        stream = BytesIO()
        file.save(stream)
        wb = load_workbook(filename=stream)
        ws = wb.active
        i = 1
        while True:
            name = ws["A"+str(i)].value
            grade = ws["B"+str(i)].value
            if not name: break

            cur.execute('INSERT INTO kinders VALUES (?,?,?)', [b64enc((name + " " + grade).encode("utf-8")).decode("utf-8"), name, grade])
            i += 1

    db.commit()
    return redirect(url_for('index'))

@app.route('/add', methods=['POST'])
def add():
    if not session.get('login'): return redirect(url_for('login'))
    if form_get('list', None) is None or form_get('list', None) == 'Выберите ученика': return h1.format('Не был выбран ученик')
    if form_get('date', None) is None: return h1.format('Не была получена дата процедуры')
    db = get_db()
    cur = db.cursor()
    form_data = form('list').split(' — ')
    proc_date = form('date').split(' — ')[1].split('.')
    uid = form_data[0] +' — '+ form_data[1]
    data = cur.execute("SELECT count FROM people WHERE uid=?", (uid, )).fetchone()
    cur.execute('UPDATE people SET count=?, last=? WHERE uid=?', [data[0]+1, ".".join(proc_date), uid])
    cur.execute('INSERT INTO proc VALUES (?,?,?,?)', (uid, proc_date[0], proc_date[1], proc_date[2]))
    db.commit()
    return redirect(url_for('index'))


@app.route('/delete', methods=['POST'])
def delete():
    if not session.get('login'): return redirect(url_for('login'))
    if form_get('delList', None) is None or form_get('delList', None) == 'Выберите ученика для удаления': return h1.format('Не был выбран ученик для удаления')
    db = get_db()
    cur = db.cursor()
    form_data = form('delList').split(' — ')
    uid = form_data[0] +' — '+ form_data[1]
    cur.execute("DELETE FROM people WHERE uid=?", (uid, ))
    db.commit()
    return redirect(url_for('index'))




@app.route('/')
def index():
    if not session.get('login'): return redirect(url_for('login'))
    db = get_db()
    cur = db.cursor()
    today = date.today()
    if session["login"] == LOGIN:
        people = cur.execute('SELECT * FROM people').fetchall()
        return render_template('index.html', people=people, str=str, today=today, timedelta=timedelta, week_days=week_days)
    else:
        kinders = cur.execute('SELECT * FROM kinders').fetchall()
        return render_template('o_index.html', kinders=kinders, b64dec=b64dec, str=str, today=today, timedelta=timedelta, week_days=week_days)




@app.route('/change', methods=['POST'])
def change():
    if not session.get('login'): return redirect(url_for('login'))
    if form_get('change', None) is None or form_get('change', None) == 'Выберите ученика для редактирования': return h1.format('Не был выбран ученик для редактирования')
    if form_get('fam', None) is None or form_get('grade', None) is None or form_get('fam', None) == '' or form_get('grade', None) == '': return h1.format('Не получено новое ФИО или Класс ученика')

    db = get_db()
    cur = db.cursor()

    old_uid = form('change')
    new_uid = cap(form("fam"))+' — '+up(form("grade"))
    if old_uid == new_uid: return h1.format("Новые данные равны старым")

    new_count = cur.execute("SELECT count, last FROM people WHERE uid=?", (old_uid, )).fetchone()
    cur_count = cur.execute("SELECT count FROM people WHERE uid=?", (new_uid, )).fetchone()
    if cur_count:
        cur_count = cur_count[0]
    else:
        cur_count = 0
        cur.execute('INSERT INTO people VALUES (?,0,?)', (new_uid, new_count[1]))

    cur.execute('UPDATE people SET count=? WHERE uid=?', [cur_count+new_count[0], new_uid])
    cur.execute('UPDATE proc SET uid=? WHERE uid=?', [new_uid, old_uid])
    cur.execute("DELETE FROM people WHERE uid=?", (old_uid, ))

    db.commit()
    return redirect(url_for('index'))




@app.route('/o_add', methods=['POST'])
def o_add():
    if not session.get('login'): return redirect(url_for('login'))
    if form_get('kinders', None) is None: return h1.format('Данные не были получены в поле kinders')
    if form_get('date', None) is None: return h1.format('Не была получена дата')
    db = get_db()
    cur = db.cursor()
    kinders = json.loads(form('kinders'))
    form_date = form('date').split(' — ')[1].split('.')
    for uid, tps in kinders.items():
        for tp in tps:
            cur.execute('INSERT INTO zan VALUES (?,?,?,?,?)', (uid, tp, form_date[0], form_date[1], form_date[2]))

    db.commit()
    return jsonify({"msg":"Успешно", "kind":0})


def o_make_xlsx(stream, when="cur"):
    xlsx = Workbook()
    db = sqlite3.connect(DB_PATH)
    cur = db.cursor()

    if when != 'cur':
        today = date.today()
        first = today.replace(day=1)
        dt = first - timedelta(days=1)
    else:
        dt = date.today()
    d, m = dt.day, dt.month
    sheet = xlsx['Sheet']
    xlsx.remove(sheet)

    sheet = xlsx.create_sheet('Отчет')
    kinders_from_db = cur.execute('SELECT uid, name, grade FROM kinders').fetchall()
    kinders = {}
    int_offset = 3
    let_offset = 3
    for i, k in enumerate(kinders_from_db, int_offset):
        kinders[k[0]] = [str(i), k[1], k[2], []]
        num = str(i)
        sheet['A'+num] = i-int_offset+1
        sheet['B'+num] = k[1]
        sheet['C'+num] = k[2]

    sheet["A"+str(int_offset-1)] = "№"
    sheet["B"+str(int_offset-1)] = "Ф.И. ребенка"
    sheet["C"+str(int_offset-1)] = "Класс"
    pre_data = {}
    last_col = let_offset
    for day in range(1, 32):
        zans = cur.execute('SELECT d, m, tp, uid FROM zan WHERE  d=? AND m=? AND y=?', (day, m, dt.year)).fetchall()
        if not zans: continue
        for z in zans:
            k = kinders[z[3]]
            call = ALPH[let_offset]+k[0]
            if not pre_data.get(call): pre_data[call] = []
            pre_data[call].append(z[2])
            k[3].append(z[2])
            if let_offset > last_col: last_col = let_offset

        sheet[ALPH[let_offset]+str(int_offset-1)] = day
        sheet[ALPH[let_offset]+str(int_offset+len(kinders.keys()))] = day
        let_offset += 1

    for cell, val in pre_data.items():
        sheet[cell] = make_count_str(val)
    for k  in kinders.values():
        sheet[ALPH[last_col+1]+k[0]] = make_count_str(k[3], "\n")
    sheet[ALPH[last_col+1]+str(int_offset-1)] = "Итог"

    xlsx.save(stream)
    xlsx.close()
    stream.seek(0)
    return stream.getvalue()

def make_count_str(data, div=""):
    s = ""
    for t, c in Counter(data).items():
        s += (str(c) + "-" + t) if c > 1 else t
        s += div
    return s

if __name__ == '__main__':
    bot_listener = Thread(target=lambda: bot.run(tele_bot), daemon=True)
    bot_db_sender = Thread(target=lambda: bot.db_auto_sender(tele_bot), daemon=True)
    bot_zip_sender = Thread(target=lambda: bot.zip_auto_sender(tele_bot), daemon=True)

    bot_listener.start()
    bot_zip_sender.start()
    bot_db_sender.start()

    app.run(host=HOST, port=PORT)
